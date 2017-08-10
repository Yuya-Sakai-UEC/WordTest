import sys
import openpyxl as xl
import random

#ファイルの読み込み

wb = xl.load_workbook('WordList.xlsx')

sheet = wb.get_sheet_by_name('Sheet1')

shsheet = wb.get_sheet_by_name('shuffleWord')

shshsheet = wb.get_sheet_by_name('shuffleWord-ans')

#スタート値とエンド値と個数の設定（今後入力型にする予定）
print("範囲と問題数を設定してください。")
print("開始番号：")
s = int(input()) #1
print("末尾番号：")
e = int(input()) #sheet.max_row
print("問題数：")
choice = int(input()) #e - s + 1

#error処理
if s<1 or s>sheet.max_row:
    print("開始番号に誤りがあります。\n")
    sys.exit(1)

if e<1 or e>sheet.max_row:
    print("末尾番号に誤りがあります。\n")
    sys.exit(1)

if s>e:
    print("たぶん開始番号と末尾番号が逆です。\n")
    sys.exit(1)

if e-s +1 < choice:
    print("問題数が足りないので全範囲を出題します。")
    choice = e - s + 1;

#乱数生成

for i in range(1,sheet.max_row+1):
    sheet.cell(row=i,column=4).value = random.random()

#乱数から指定した区間の部分をソート

a = []

for i in range(s,e+1):
    a = a + [[sheet.cell(row=i,column=1).value,sheet.cell(row=i,column=2).value,sheet.cell(row=i,column=3).value,sheet.cell(row=i,column=4).value]]

# sorted(object or list,key=lambda object:object.choice)
b = sorted(a,key=lambda a:a[3])


#shuffleWord,shuffleWord-ansにデータを格納

for i in range(choice):
    shsheet.cell(row=i+2,column=1).value = b[i][0]
    shsheet.cell(row=i+2,column=2).value = b[i][1]

for i in range(choice):
    shshsheet.cell(row=i+2,column=1).value = b[i][0]
    shshsheet.cell(row=i+2,column=2).value = b[i][1]
    shshsheet.cell(row=i+2,column=3).value = b[i][2]

#save

wb.save('WordList.xlsx')
